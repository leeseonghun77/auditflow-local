from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

import fill_ledger_from_pdfs as fill

BASE = Path.cwd()
INPUT = BASE / "26-1_감사자동채움.xlsx"
OUTPUT = BASE / "26-1_감사기준보완.xlsx"
EVIDENCE_PDF = BASE / "26-1학기 TUBE 지출증빙자료.pdf"

@dataclass
class Evidence:
    number: int
    title: str
    date: datetime
    note: str


def yymmdd_to_date(text: str) -> datetime:
    y, m, d = [int(x) for x in text.split(".")]
    return datetime(2000 + y, m, d)


def read_evidence() -> list[Evidence]:
    reader = PdfReader(str(EVIDENCE_PDF))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    pattern = re.compile(
        r"번호\s*(?P<number>\d+)\s*항목명\s*(?P<title>.*?)\s*날짜\s*(?P<date>\d{2}\.\d{2}\.\d{2})(?P<tail>.*?)(?=번호\s*\d+\s*항목명|\Z)",
        re.S,
    )
    rows: list[Evidence] = []
    for match in pattern.finditer(text):
        title = re.sub(r"\s+", " ", match.group("title")).strip()
        tail = re.sub(r"\s+", " ", match.group("tail")).strip()
        note = ""
        if "비고" in tail:
            note = tail.split("비고", 1)[1].strip()[:120]
        rows.append(Evidence(int(match.group("number")), title, yymmdd_to_date(match.group("date")), note))
    return rows


def read_transactions() -> list[fill.Transaction]:
    txs: list[fill.Transaction] = []
    txs.extend(fill.parse_kakao(BASE / "2026-01 TUBE 거래내역서 1(카카오뱅크).pdf"))
    txs.extend(fill.parse_toss(BASE / "2026-01 TUBE 거래내역서 2(토스뱅크).pdf"))
    txs.sort(key=lambda tx: tx.dt)
    return txs


def classify(tx: fill.Transaction) -> tuple[str, str]:
    text = f"{tx.name} {tx.detail}"
    if tx.amount < 0:
        if "출금" in tx.detail or "일반이체" in tx.detail:
            return "감사대상 지출", "계좌이체 지출입니다. 이체확인증 또는 사유 확인이 필요합니다."
        if "체크카드" in tx.detail:
            return "감사대상 지출", "체크카드 지출입니다. 지출증빙자료와 대조하세요."
        return "감사대상 지출", "출금 거래입니다. 증빙 및 지출 사유 확인이 필요합니다."
    if "캐시백" in text:
        return "입금-캐시백", "카드 캐시백입니다. 원 지출과 함께 설명하면 안전합니다."
    if "이자" in text:
        return "입금-이자", "통장 이자입니다. 거래내역만으로 성격 확인 가능합니다."
    if "잔액 이체" in text:
        return "입금-잔액이체", "계좌 이전으로 보입니다. 이전 계좌 마지막 잔액과 연결 설명이 필요합니다."
    return "입금", "회비/참가비/환급/오입금 여부를 필요 시 비고에 보완하세요."


def score_evidence(tx_no: int, tx: fill.Transaction, ev: Evidence) -> int:
    score = 0
    day_diff = abs((tx.dt.date() - ev.date.date()).days)
    if ev.number == tx_no:
        score += 45
    elif abs(ev.number - tx_no) == 1:
        score += 20
    if day_diff == 0:
        score += 35
    elif day_diff == 1:
        score += 20
    elif day_diff <= 3:
        score += 8
    words = set(re.findall(r"[가-힣A-Za-z0-9㈜]{2,}", tx.name + " " + tx.detail))
    ev_words = set(re.findall(r"[가-힣A-Za-z0-9㈜]{2,}", ev.title + " " + ev.note))
    common = words & ev_words
    score += min(20, len(common) * 5)
    if "오입금" in ev.title and tx.amount < 0:
        score += 20
    if "오지출" in ev.title:
        score += 10
    return score


def find_evidence(tx_no: int, tx: fill.Transaction, evidences: list[Evidence]) -> tuple[Evidence | None, int, str]:
    if tx.amount >= 0 and "오입금" not in tx.name + tx.detail:
        return None, 0, "입금 거래라 증빙 매칭 대상 아님"
    scored = sorted(((score_evidence(tx_no, tx, ev), ev) for ev in evidences), key=lambda item: item[0], reverse=True)
    if not scored or scored[0][0] < 30:
        return None, 0, "증빙 후보 없음 - 수기 확인 필요"
    score, ev = scored[0]
    if ev.number == tx_no:
        reason = "증빙번호와 엑셀번호가 일치"
    elif abs(ev.number - tx_no) <= 1:
        reason = "증빙번호가 인접하고 날짜/내용이 유사"
    else:
        reason = "날짜/내용 기준 후보"
    return ev, score, reason


def style_header(ws, row: int, start_col: int, end_col: int) -> None:
    fill_color = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill_color
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def enhance() -> None:
    if not INPUT.exists():
        fill.fill_workbook(read_transactions())
    txs = read_transactions()
    evidences = read_evidence()

    wb = load_workbook(INPUT)
    ws = wb["2026-01"]

    headers = {
        8: "감사구분",
        9: "증빙번호",
        10: "증빙항목",
        11: "감사비고",
    }
    for col, value in headers.items():
        ws.cell(3, col).value = value
    style_header(ws, 3, 8, 11)

    start_row = 4
    for idx, tx in enumerate(txs, 1):
        row = start_row + idx - 1
        ws.cell(row, 1).value = idx
        category, base_note = classify(tx)
        ev, score, reason = find_evidence(idx, tx, evidences)
        ws.cell(row, 8).value = category
        if ev:
            ws.cell(row, 9).value = ev.number
            ws.cell(row, 10).value = ev.title
            note_parts = [reason]
            if ev.number != idx:
                note_parts.append(f"증빙번호({ev.number})와 엑셀번호({idx}) 불일치 - 제출 전 확인")
            if abs((tx.dt.date() - ev.date.date()).days) > 1:
                note_parts.append(f"증빙일({ev.date.strftime('%Y-%m-%d')})과 거래일({tx.dt.strftime('%Y-%m-%d')}) 차이 확인")
            if "오입금" in ev.title or "오지출" in ev.title:
                note_parts.append("오입금/오지출 관련 항목으로 기록 필요")
            note_parts.append(base_note)
            ws.cell(row, 11).value = " / ".join(note_parts)
        else:
            ws.cell(row, 9).value = ""
            ws.cell(row, 10).value = ""
            ws.cell(row, 11).value = base_note if tx.amount >= 0 else f"{reason} / {base_note}"

        for col in range(8, 12):
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)

    # Clear numbering in unused rows so submission sheet does not imply phantom entries.
    for row in range(start_row + len(txs), ws.max_row + 1):
        for col in range(1, 12):
            ws.cell(row, col).value = None

    widths = {1: 8, 2: 12, 3: 18, 4: 26, 5: 12, 6: 12, 7: 14, 8: 16, 9: 10, 10: 38, 11: 55}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"

    if "감사보완_증빙목록" in wb.sheetnames:
        del wb["감사보완_증빙목록"]
    ev_ws = wb.create_sheet("감사보완_증빙목록")
    ev_ws.append(["증빙번호", "증빙일", "증빙항목", "비고"])
    for ev in evidences:
        ev_ws.append([ev.number, ev.date.date(), ev.title, ev.note])
    style_header(ev_ws, 1, 1, 4)
    for row in ev_ws.iter_rows(min_row=2):
        row[1].number_format = "yyyy-mm-dd"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in enumerate([10, 14, 48, 60], 1):
        ev_ws.column_dimensions[get_column_letter(col)].width = width
    ev_ws.freeze_panes = "A2"

    if "감사보완_요약" in wb.sheetnames:
        del wb["감사보완_요약"]
    sum_ws = wb.create_sheet("감사보완_요약")
    withdrawals = [tx for tx in txs if tx.amount < 0]
    matched = 0
    mismatch = 0
    for idx, tx in enumerate(txs, 1):
        ev, _, _ = find_evidence(idx, tx, evidences)
        if tx.amount < 0 and ev:
            matched += 1
            if ev.number != idx:
                mismatch += 1
    rows = [
        ["항목", "값"],
        ["전체 거래 수", len(txs)],
        ["출금 거래 수", len(withdrawals)],
        ["증빙 후보 연결 출금 수", matched],
        ["증빙번호와 엑셀번호 불일치 후보", mismatch],
        ["총 입금", sum(tx.amount for tx in txs if tx.amount > 0)],
        ["총 출금", sum(abs(tx.amount) for tx in txs if tx.amount < 0)],
        ["보완 방식", "A열을 1부터 재번호화하고 H~K열에 감사구분/증빙번호/증빙항목/비고를 추가"],
    ]
    for row in rows:
        sum_ws.append(row)
    style_header(sum_ws, 1, 1, 2)
    sum_ws.column_dimensions["A"].width = 32
    sum_ws.column_dimensions["B"].width = 80

    wb.save(OUTPUT)
    print(f"created={OUTPUT}")
    print(f"transactions={len(txs)}")
    print(f"evidences={len(evidences)}")
    print(f"withdrawals={len(withdrawals)}")
    print(f"matched_withdrawals={matched}")
    print(f"number_mismatch_candidates={mismatch}")


if __name__ == "__main__":
    enhance()