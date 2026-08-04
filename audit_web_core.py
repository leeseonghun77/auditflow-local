from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

import fill_ledger_from_pdfs as fill


@dataclass
class Evidence:
    number: int
    title: str
    date: datetime | None
    note: str


@dataclass
class BuildResult:
    output_path: Path
    transaction_count: int
    withdrawal_count: int
    evidence_count: int
    matched_withdrawals: int
    mismatch_candidates: int
    deposit_sum: int
    withdrawal_sum: int
    warnings: list[str]


def read_pdf_text(path: Path, max_pages: int | None = None) -> str:
    reader = PdfReader(str(path))
    pages = reader.pages if max_pages is None else reader.pages[:max_pages]
    return "\n".join(page.extract_text() or "" for page in pages)


def detect_bank_pdf(path: Path) -> str:
    text = read_pdf_text(path, max_pages=1)
    if "카카오" in text or re.search(r"\d{4}\.\d{2}\.\d{2} \d{2}:\d{2}:\d{2}", text):
        return "kakao"
    if "토스" in text or re.search(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}", text):
        return "toss"
    return "unknown"


def parse_bank_files(paths: Iterable[Path]) -> tuple[list[fill.Transaction], list[str]]:
    transactions: list[fill.Transaction] = []
    warnings: list[str] = []
    for path in paths:
        kind = detect_bank_pdf(path)
        try:
            if kind == "kakao":
                parsed = fill.parse_kakao(path)
            elif kind == "toss":
                parsed = fill.parse_toss(path)
            else:
                warnings.append(f"은행 PDF 형식을 판별하지 못했습니다: {path.name}")
                continue
            if not parsed:
                warnings.append(f"거래를 추출하지 못했습니다: {path.name}")
            transactions.extend(parsed)
        except Exception as exc:
            warnings.append(f"{path.name} 처리 실패: {type(exc).__name__}: {exc}")
    transactions.sort(key=lambda tx: tx.dt)
    return transactions, warnings


def parse_yymmdd(text: str) -> datetime | None:
    try:
        y, m, d = [int(x) for x in text.split(".")]
        return datetime(2000 + y, m, d)
    except Exception:
        return None


def parse_evidence_pdf(path: Path | None) -> tuple[list[Evidence], list[str]]:
    if not path:
        return [], []
    warnings: list[str] = []
    try:
        text = read_pdf_text(path)
    except Exception as exc:
        return [], [f"증빙 PDF를 읽지 못했습니다: {type(exc).__name__}: {exc}"]

    pattern = re.compile(
        r"번호\s*(?P<number>\d+)\s*항목명\s*(?P<title>.*?)\s*날짜\s*(?P<date>\d{2}\.\d{2}\.\d{2})(?P<tail>.*?)(?=번호\s*\d+\s*항목명|\Z)",
        re.S,
    )
    evidences: list[Evidence] = []
    for match in pattern.finditer(text):
        title = re.sub(r"\s+", " ", match.group("title")).strip()
        tail = re.sub(r"\s+", " ", match.group("tail")).strip()
        note = ""
        if "비고" in tail:
            note = tail.split("비고", 1)[1].strip()[:140]
        evidences.append(Evidence(int(match.group("number")), title, parse_yymmdd(match.group("date")), note))
    if not evidences:
        warnings.append("증빙 PDF에서 '번호/항목명/날짜' 구조를 찾지 못했습니다.")
    return evidences, warnings


def classify(tx: fill.Transaction) -> tuple[str, str]:
    text = f"{tx.name} {tx.detail}"
    if tx.amount < 0:
        if "출금" in tx.detail or "일반이체" in tx.detail:
            return "감사대상 지출", "계좌이체 지출입니다. 이체확인증 또는 사유 확인이 필요합니다."
        if "체크카드" in tx.detail:
            return "감사대상 지출", "체크카드 지출입니다. 지출증빙자료와 대조하세요."
        return "감사대상 지출", "출금 거래입니다. 증빙 및 지출 사유 확인이 필요합니다."
    if "캐시백" in text:
        return "입금-캐시백", "카드 캐시백입니다. 원 지출과 함께 설명하면 안전합니다."
    if "이자" in text:
        return "입금-이자", "통장 이자입니다. 거래내역만으로 성격 확인 가능합니다."
    if "잔액 이체" in text:
        return "입금-잔액이체", "계좌 이전으로 보입니다. 이전 계좌 마지막 잔액과 연결 설명이 필요합니다."
    return "입금", "회비/참가비/환급/오입금 여부를 필요 시 비고에 보완하세요."


def score_evidence(tx_no: int, tx: fill.Transaction, ev: Evidence) -> int:
    score = 0
    if ev.number == tx_no:
        score += 45
    elif abs(ev.number - tx_no) == 1:
        score += 20
    if ev.date:
        day_diff = abs((tx.dt.date() - ev.date.date()).days)
        if day_diff == 0:
            score += 35
        elif day_diff == 1:
            score += 20
        elif day_diff <= 3:
            score += 8
    words = set(re.findall(r"[가-힣A-Za-z0-9㈜]{2,}", tx.name + " " + tx.detail))
    ev_words = set(re.findall(r"[가-힣A-Za-z0-9㈜]{2,}", ev.title + " " + ev.note))
    score += min(20, len(words & ev_words) * 5)
    if "오입금" in ev.title and tx.amount < 0:
        score += 20
    if "오지출" in ev.title:
        score += 10
    return score


def find_evidence(tx_no: int, tx: fill.Transaction, evidences: list[Evidence]) -> tuple[Evidence | None, int, str]:
    if not evidences:
        return None, 0, "증빙 PDF 없음"
    if tx.amount >= 0:
        return None, 0, "입금 거래라 증빙 매칭 대상 아님"
    scored = sorted(((score_evidence(tx_no, tx, ev), ev) for ev in evidences), key=lambda item: item[0], reverse=True)
    if not scored or scored[0][0] < 30:
        return None, 0, "증빙 후보 없음 - 수기 확인 필요"
    score, ev = scored[0]
    if ev.number == tx_no:
        reason = "증빙번호와 엑셀번호가 일치"
    elif abs(ev.number - tx_no) <= 1:
        reason = "증빙번호가 인접하고 날짜/내용이 유사"
    else:
        reason = "날짜/내용 기준 후보"
    return ev, score, reason


def style_header(ws, row: int, start_col: int, end_col: int) -> None:
    fill_color = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill_color
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def copy_basic_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, 12):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)


def build_workbook(
    template_path: Path,
    bank_pdf_paths: list[Path],
    evidence_pdf_path: Path | None,
    output_path: Path,
    *,
    start_number: int = 1,
    include_helper_sheets: bool = True,
) -> BuildResult:
    transactions, warnings = parse_bank_files(bank_pdf_paths)
    evidences, evidence_warnings = parse_evidence_pdf(evidence_pdf_path)
    warnings.extend(evidence_warnings)
    if not transactions:
        raise ValueError("은행 거래내역 PDF에서 거래를 추출하지 못했습니다.")

    wb = load_workbook(template_path)
    ws = wb["2026-01"] if "2026-01" in wb.sheetnames else wb.active
    start_row = 4
    needed_last_row = start_row + len(transactions) - 1
    if ws.max_row < needed_last_row:
        ws.insert_rows(ws.max_row + 1, needed_last_row - ws.max_row)

    headers = {8: "감사구분", 9: "증빙번호", 10: "증빙항목", 11: "감사비고"}
    for col, value in headers.items():
        ws.cell(3, col).value = value
    style_header(ws, 3, 8, 11)

    for row in range(start_row, ws.max_row + 1):
        for col in range(1, 12):
            ws.cell(row, col).value = None

    matched_withdrawals = 0
    mismatch_candidates = 0
    for idx, tx in enumerate(transactions, start_number):
        row = start_row + idx - start_number
        if row > start_row:
            copy_basic_style(ws, start_row, row)
        tx_no = idx
        ws.cell(row, 1).value = tx_no
        ws.cell(row, 2).value = tx.dt.date()
        ws.cell(row, 3).value = tx.name
        ws.cell(row, 4).value = f"{tx.bank} {tx.detail}"
        ws.cell(row, 5).value = tx.amount if tx.amount > 0 else None
        ws.cell(row, 6).value = abs(tx.amount) if tx.amount < 0 else None
        ws.cell(row, 7).value = tx.balance
        ws.cell(row, 2).number_format = "yyyy-mm-dd"
        for col in (5, 6, 7):
            ws.cell(row, col).number_format = "#,##0"

        category, base_note = classify(tx)
        ev, _, reason = find_evidence(tx_no, tx, evidences)
        ws.cell(row, 8).value = category
        if ev:
            if tx.amount < 0:
                matched_withdrawals += 1
            ws.cell(row, 9).value = ev.number
            ws.cell(row, 10).value = ev.title
            notes = [reason]
            if ev.number != tx_no:
                mismatch_candidates += 1
                notes.append(f"증빙번호({ev.number})와 엑셀번호({tx_no}) 불일치 - 제출 전 확인")
            if ev.date and abs((tx.dt.date() - ev.date.date()).days) > 1:
                notes.append(f"증빙일({ev.date.strftime('%Y-%m-%d')})과 거래일({tx.dt.strftime('%Y-%m-%d')}) 차이 확인")
            if "오입금" in ev.title or "오지출" in ev.title:
                notes.append("오입금/오지출 관련 항목으로 기록 필요")
            notes.append(base_note)
            ws.cell(row, 11).value = " / ".join(notes)
        else:
            ws.cell(row, 11).value = base_note if tx.amount >= 0 else f"{reason} / {base_note}"

        for col in range(1, 12):
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)

    for col, width in {1: 8, 2: 12, 3: 18, 4: 26, 5: 12, 6: 12, 7: 14, 8: 16, 9: 10, 10: 38, 11: 55}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"

    if include_helper_sheets:
        for sheet_name in ["자동채움_원문", "감사보완_증빙목록", "감사보완_요약"]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        raw_ws = wb.create_sheet("자동채움_원문")
        raw_ws.append(["순번", "은행", "거래일시", "구분", "금액", "거래후잔액", "상대/가맹점", "거래구분", "PDF 원문"])
        for i, tx in enumerate(transactions, start_number):
            raw_ws.append([i, tx.bank, tx.dt.strftime("%Y-%m-%d %H:%M:%S"), tx.kind, tx.amount, tx.balance, tx.name, tx.detail, tx.raw])
        style_header(raw_ws, 1, 1, 9)
        raw_ws.freeze_panes = "A2"
        for col, width in enumerate([8, 12, 22, 8, 12, 14, 28, 18, 80], 1):
            raw_ws.column_dimensions[get_column_letter(col)].width = width

        ev_ws = wb.create_sheet("감사보완_증빙목록")
        ev_ws.append(["증빙번호", "증빙일", "증빙항목", "비고"])
        for ev in evidences:
            ev_ws.append([ev.number, ev.date.date() if ev.date else "", ev.title, ev.note])
        style_header(ev_ws, 1, 1, 4)
        ev_ws.freeze_panes = "A2"
        for col, width in enumerate([10, 14, 48, 60], 1):
            ev_ws.column_dimensions[get_column_letter(col)].width = width

        sum_ws = wb.create_sheet("감사보완_요약")
        withdrawals = [tx for tx in transactions if tx.amount < 0]
        rows = [
            ["항목", "값"],
            ["전체 거래 수", len(transactions)],
            ["출금 거래 수", len(withdrawals)],
            ["증빙 후보 연결 출금 수", matched_withdrawals],
            ["증빙번호와 엑셀번호 불일치 후보", mismatch_candidates],
            ["총 입금", sum(tx.amount for tx in transactions if tx.amount > 0)],
            ["총 출금", sum(abs(tx.amount) for tx in transactions if tx.amount < 0)],
        ]
        for row in rows:
            sum_ws.append(row)
        style_header(sum_ws, 1, 1, 2)
        sum_ws.column_dimensions["A"].width = 32
        sum_ws.column_dimensions["B"].width = 72

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return BuildResult(
        output_path=output_path,
        transaction_count=len(transactions),
        withdrawal_count=sum(1 for tx in transactions if tx.amount < 0),
        evidence_count=len(evidences),
        matched_withdrawals=matched_withdrawals,
        mismatch_candidates=mismatch_candidates,
        deposit_sum=sum(tx.amount for tx in transactions if tx.amount > 0),
        withdrawal_sum=sum(abs(tx.amount) for tx in transactions if tx.amount < 0),
        warnings=warnings,
    )