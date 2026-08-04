from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pypdf import PdfReader

BASE = Path.cwd()
SOURCE_XLSX = BASE / "2026-01 TUBE 회계(26.05.26~26.07.17).xlsx"
OUTPUT_XLSX = BASE / "26-1_감사자동채움.xlsx"

@dataclass
class Transaction:
    dt: datetime
    bank: str
    kind: str
    amount: int
    balance: int
    name: str
    detail: str
    raw: str

    @property
    def deposit(self) -> int | None:
        return self.amount if self.amount > 0 else None

    @property
    def withdrawal(self) -> int | None:
        return abs(self.amount) if self.amount < 0 else None


def money(text: str) -> int:
    return int(text.replace(",", ""))


def pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def parse_kakao(path: Path) -> list[Transaction]:
    text = pdf_text(path)
    txs: list[Transaction] = []
    kinds = ["일반이체", "예금이자", "일반입금", "체크카드결제", "캐시백"]
    kind_re = "|".join(map(re.escape, kinds))
    pattern = re.compile(
        rf"^(?P<date>\d{{4}}\.\d{{2}}\.\d{{2}}) (?P<time>\d{{2}}:\d{{2}}:\d{{2}}) "
        rf"(?P<direction>입금|출금) (?P<amount>-?[\d,]+) (?P<balance>[\d,]+) "
        rf"(?P<name>.+?) (?P<detail>{kind_re})$"
    )
    for raw in text.splitlines():
        line = re.sub(r"\s+", " ", raw).strip()
        m = pattern.match(line)
        if not m:
            continue
        amount = money(m.group("amount"))
        if m.group("direction") == "출금" and amount > 0:
            amount = -amount
        txs.append(
            Transaction(
                dt=datetime.strptime(m.group("date") + " " + m.group("time"), "%Y.%m.%d %H:%M:%S"),
                bank="카카오뱅크",
                kind=m.group("direction"),
                amount=amount,
                balance=money(m.group("balance")),
                name=m.group("name").strip(),
                detail=m.group("detail").strip(),
                raw=line,
            )
        )
    return txs


def parse_toss(path: Path) -> list[Transaction]:
    text = pdf_text(path)
    txs: list[Transaction] = []
    pattern = re.compile(
        r"^(?P<date>\d{4}-\d{2}-\d{2}) (?P<time>\d{2}:\d{2}:\d{2}) "
        r"(?P<detail>이자입금|체크카드결제|입금|출금) (?P<amount>-?[\d,]+) (?P<balance>[\d,]+) (?P<name>.+)$"
    )
    for raw in text.splitlines():
        line = re.sub(r"\s+", " ", raw).strip()
        m = pattern.match(line)
        if not m:
            continue
        amount = money(m.group("amount"))
        if m.group("detail") in {"출금", "체크카드결제"} and amount > 0:
            amount = -amount
        kind = "입금" if amount > 0 else "출금"
        txs.append(
            Transaction(
                dt=datetime.strptime(m.group("date") + " " + m.group("time"), "%Y-%m-%d %H:%M:%S"),
                bank="토스뱅크",
                kind=kind,
                amount=amount,
                balance=money(m.group("balance")),
                name=m.group("name").strip(),
                detail=m.group("detail").strip(),
                raw=line,
            )
        )
    return txs


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, 12):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def fill_workbook(transactions: list[Transaction]) -> None:
    wb = load_workbook(SOURCE_XLSX)
    ws = wb["2026-01"] if "2026-01" in wb.sheetnames else wb.active
    start_row = 4

    # 기존 양식의 A열 연번은 보존하고, 거래 입력 영역만 초기화합니다.
    for row in range(start_row, ws.max_row + 1):
        for col in range(2, 8):
            ws.cell(row, col).value = None

    needed_last_row = start_row + len(transactions) - 1
    if ws.max_row < needed_last_row:
        ws.insert_rows(ws.max_row + 1, needed_last_row - ws.max_row)

    # A열 연번이 모자라면 기존 패턴을 이어갑니다.
    last_no = None
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, int):
            last_no = value
    for row in range(start_row, needed_last_row + 1):
        if ws.cell(row, 1).value is None:
            last_no = 1 if last_no is None else last_no + 1
            ws.cell(row, 1).value = last_no
            copy_row_style(ws, start_row, row)

    for offset, tx in enumerate(transactions):
        row = start_row + offset
        ws.cell(row, 2).value = tx.dt.date()
        ws.cell(row, 3).value = tx.name
        ws.cell(row, 4).value = f"{tx.bank} {tx.detail}"
        ws.cell(row, 5).value = tx.deposit
        ws.cell(row, 6).value = tx.withdrawal
        ws.cell(row, 7).value = tx.balance
        ws.cell(row, 2).number_format = "yyyy-mm-dd"
        for col in (5, 6, 7):
            ws.cell(row, col).number_format = "#,##0"
        for col in range(2, 8):
            ws.cell(row, col).alignment = Alignment(horizontal="center" if col != 4 else "left", vertical="center", wrap_text=True)

    # 남는 A열 연번 행은 그대로 두되 빈 입력 영역으로 유지합니다.
    if "자동채움_원문" in wb.sheetnames:
        del wb["자동채움_원문"]
    raw_ws = wb.create_sheet("자동채움_원문")
    raw_ws.append(["순번", "은행", "거래일시", "구분", "금액", "거래후잔액", "상대/가맹점", "거래구분", "PDF 원문"])
    for i, tx in enumerate(transactions, 1):
        raw_ws.append([i, tx.bank, tx.dt.strftime("%Y-%m-%d %H:%M:%S"), tx.kind, tx.amount, tx.balance, tx.name, tx.detail, tx.raw])
    for cell in raw_ws[1]:
        cell.style = "Headline 4"
    raw_ws.freeze_panes = "A2"
    widths = [8, 12, 22, 8, 12, 14, 28, 18, 80]
    for idx, width in enumerate(widths, 1):
        raw_ws.column_dimensions[raw_ws.cell(1, idx).column_letter].width = width

    wb.save(OUTPUT_XLSX)


def main() -> None:
    transactions: list[Transaction] = []
    transactions.extend(parse_kakao(BASE / "2026-01 TUBE 거래내역서 1(카카오뱅크).pdf"))
    transactions.extend(parse_toss(BASE / "2026-01 TUBE 거래내역서 2(토스뱅크).pdf"))
    transactions.sort(key=lambda tx: tx.dt)
    fill_workbook(transactions)
    print(f"filled={OUTPUT_XLSX}")
    print(f"transactions={len(transactions)}")
    print(f"first={transactions[0].dt if transactions else ''}")
    print(f"last={transactions[-1].dt if transactions else ''}")


if __name__ == "__main__":
    main()